import tkinter as tk
from tkinter import messagebox, ttk
import mysql.connector
import requests
import os
import uuid
from dotenv import load_dotenv
from datetime import datetime
from fpdf import FPDF
from openpyxl import Workbook

load_dotenv()

DB_CONFIG = {
    "user": os.environ["DB_USER"],
    "password": os.environ["DB_PASS"],
    "host": os.environ["DB_HOST"],
    "database": os.environ["DB_NAME"],
}

DICTIONARY_API_URL = "https://api.dictionaryapi.dev/api/v2/entries/en/"

TIMESTAMP = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")


def generate_machine_id():
    """Generate a unique identifier for the current machine."""
    return str(uuid.UUID(int=uuid.getnode()))


MACHINE_ID = generate_machine_id()


def init_db():
    """Initialize the database schema with the required tables."""
    conn = mysql.connector.connect(**DB_CONFIG)
    cursor = conn.cursor()

    # Create vocabulary table
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS vocabulary (
            id INT AUTO_INCREMENT PRIMARY KEY,
            word VARCHAR(255) NOT NULL,
            meaning TEXT NOT NULL,
            UNIQUE(word)
        )
    """
    )

    # Create license_keys table with machine_id feature and max_machines
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS license_keys (
            key_id INT AUTO_INCREMENT PRIMARY KEY,
            license_key VARCHAR(255) NOT NULL UNIQUE,
            max_machines INT NOT NULL DEFAULT 1,
            status ENUM('active', 'revoked') NOT NULL DEFAULT 'active',
            expiry_date DATE DEFAULT NULL
        )
    """
    )

    # Create machine_activations table
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS machine_activations (
            activation_id INT AUTO_INCREMENT PRIMARY KEY,
            license_key VARCHAR(255) NOT NULL,
            machine_id VARCHAR(255) NOT NULL,
            activation_date DATETIME DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(license_key, machine_id),
            FOREIGN KEY (license_key) REFERENCES license_keys(license_key)
        )
    """
    )

    conn.commit()
    conn.close()


# Global variable to track the open definition window
definition_window = None


def validate_license_key(license_key):
    """Validate the license key and ensure it matches the machine."""
    conn = mysql.connector.connect(**DB_CONFIG)
    cursor = conn.cursor()

    # Check if the current machine is already activated
    cursor.execute(
        "SELECT activation_id FROM machine_activations WHERE license_key = %s AND machine_id = %s",
        (license_key, MACHINE_ID),
    )
    existing_activation = cursor.fetchone()

    if existing_activation:
        conn.close()
        return True, "Machine is already activated with this license key."

    # Check license key status and max_machines
    cursor.execute(
        "SELECT status, expiry_date, max_machines FROM license_keys WHERE license_key = %s",
        (license_key,),
    )
    result = cursor.fetchone()

    if not result:
        conn.close()
        return False, "Invalid license key."

    status, expiry_date, max_machines = result

    # Check status
    if status != "active":
        conn.close()
        return False, "License key is not active."

    # Check expiry
    if expiry_date and datetime.strptime(expiry_date, "%Y-%m-%d") < datetime.now():
        conn.close()
        return False, "License key has expired."

    # Check number of activated machines
    cursor.execute(
        "SELECT COUNT(*) FROM machine_activations WHERE license_key = %s",
        (license_key,),
    )
    activated_machines = cursor.fetchone()[0]

    if activated_machines >= max_machines:
        conn.close()
        return (
            False,
            f"Maximum number of machines ({max_machines}) already activated for this license key.",
        )

    # Activate the machine
    try:
        cursor.execute(
            "INSERT INTO machine_activations (license_key, machine_id) VALUES (%s, %s)",
            (license_key, MACHINE_ID),
        )
        conn.commit()
    except mysql.connector.IntegrityError:
        conn.close()
        return False, "An error occurred while activating the machine."

    conn.close()
    return True, "License key validated successfully and machine activated."


def fetch_meaning(word):
    """Fetch the meaning of a word using the dictionary API."""
    url = f"{DICTIONARY_API_URL}{word.lower()}"
    response = requests.get(url)
    if response.status_code != 200:
        return None
    data = response.json()
    return data[0]["meanings"][0]["definitions"][0]["definition"]


def add_word(event=None):
    """Add a new word and its meaning to the database."""
    word = entry_word.get().strip()
    meaning = entry_meaning.get().strip()

    if not word:
        messagebox.showwarning("Input Error", "Please provide a word.")
        return

    conn = mysql.connector.connect(**DB_CONFIG)
    cursor = conn.cursor()

    # Check if the word already exists
    cursor.execute("SELECT COUNT(*) FROM vocabulary WHERE word = %s", (word,))
    exists = cursor.fetchone()[0]

    if exists:
        conn.close()
        messagebox.showerror(
            "Error", f"The word '{word}' already exists in the database."
        )
        return

    if not meaning:
        set_cursor("wait")  # Show loading cursor
        meaning = fetch_meaning(word)
        set_cursor("")  # Reset cursor
        if not meaning:
            messagebox.showwarning(
                "Fetch Error", "Could not fetch meaning from the dictionary."
            )
            conn.close()
            return

    try:
        cursor.execute(
            "INSERT INTO vocabulary (word, meaning) VALUES (%s, %s)", (word, meaning)
        )
        conn.commit()
        conn.close()
        messagebox.showinfo("Success", f"'{word}' added successfully!")
        entry_word.delete(0, tk.END)
        entry_meaning.delete(0, tk.END)
        load_vocabulary()
    except mysql.connector.Error as e:
        conn.close()
        messagebox.showerror("Database Error", f"An error occurred: {e}")


def edit_word():
    """Edit the selected word and update the database."""
    selected_item = tree_vocabulary.selection()
    if not selected_item:
        messagebox.showwarning("Selection Error", "Please select a word to edit.")
        return

    word = tree_vocabulary.item(selected_item, "values")[0]

    # Prompt user for new word and meaning
    edit_window = tk.Toplevel(root)
    edit_window.title("Edit Word")
    edit_window.geometry("400x200")

    ttk.Label(edit_window, text="Word/Phrase:").pack(pady=5)
    entry_new_word = ttk.Entry(edit_window, font=("Verdana", 12))
    entry_new_word.pack(pady=5, padx=10, fill=tk.X)
    entry_new_word.insert(0, word)

    ttk.Label(edit_window, text="Meaning:").pack(pady=5)
    entry_new_meaning = ttk.Entry(edit_window, font=("Verdana", 12))
    entry_new_meaning.pack(pady=5, padx=10, fill=tk.X)

    conn = mysql.connector.connect(**DB_CONFIG)
    cursor = conn.cursor()
    cursor.execute("SELECT meaning FROM vocabulary WHERE word = %s", (word,))
    old_meaning = cursor.fetchone()
    if old_meaning:
        entry_new_meaning.insert(0, old_meaning[0])
    conn.close()

    def save_changes():
        new_word = entry_new_word.get().strip()
        new_meaning = entry_new_meaning.get().strip()

        if not new_word or not new_meaning:
            messagebox.showwarning(
                "Input Error", "Both Word and Meaning fields must be filled."
            )
            return

        conn = mysql.connector.connect(**DB_CONFIG)
        cursor = conn.cursor()
        try:
            cursor.execute(
                "UPDATE vocabulary SET word = %s, meaning = %s WHERE word = %s",
                (new_word, new_meaning, word),
            )
            conn.commit()
            conn.close()
            messagebox.showinfo("Success", "Word updated successfully!")
            edit_window.destroy()
            load_vocabulary()
        except mysql.connector.IntegrityError:
            conn.close()
            messagebox.showerror("Error", "This word already exists in the database.")

    ttk.Button(edit_window, text="Save Changes", command=save_changes).pack(pady=10)

    edit_window.protocol("WM_DELETE_WINDOW", edit_window.destroy)


def delete_word():
    selected_item = tree_vocabulary.selection()
    if not selected_item:
        messagebox.showwarning("Selection Error", "Please select a word to delete.")
        return

    word = tree_vocabulary.item(selected_item, "values")[0]
    conn = mysql.connector.connect(**DB_CONFIG)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM vocabulary WHERE word = %s", (word,))
    conn.commit()
    conn.close()

    messagebox.showinfo("Success", "Word deleted successfully.")
    load_vocabulary()


def on_search():
    search_term = entry_search.get().strip()
    load_vocabulary(search_term)


def load_vocabulary(search_term=""):
    """Load vocabulary words and meanings into the Treeview."""
    conn = mysql.connector.connect(**DB_CONFIG)
    cursor = conn.cursor()
    query = "SELECT word, meaning FROM vocabulary"
    if search_term:
        query += " WHERE word LIKE %s"
        cursor.execute(query, (f"%{search_term}%",))
    else:
        cursor.execute(query)
    words = cursor.fetchall()
    conn.close()

    for row in tree_vocabulary.get_children():
        tree_vocabulary.delete(row)

    for word in words:
        tree_vocabulary.insert("", "end", values=word)


def set_cursor(cursor_type):
    """Set the cursor type for the application."""
    root.config(cursor=cursor_type)
    root.update_idletasks()


def show_license_key_entry():
    """Prompt user to enter a license key for validation or skip if already validated."""
    conn = mysql.connector.connect(**DB_CONFIG)
    cursor = conn.cursor()

    # Check if the machine is already activated
    cursor.execute(
        "SELECT license_key FROM machine_activations WHERE machine_id = %s",
        (MACHINE_ID,),
    )
    result = cursor.fetchone()
    conn.close()

    if result:
        # If the machine is already activated, skip the license key input
        root.deiconify()  # Show the main application window
        return

    # Prompt for license key if the machine is not activated
    def submit_key():
        license_key = entry_license.get().strip()
        valid, message = validate_license_key(license_key)

        if valid:
            messagebox.showinfo("License Validation", message)
            license_window.destroy()
            root.deiconify()  # Show the main application window
        else:
            messagebox.showerror("License Validation", message)

    root.withdraw()  # Hide the main application window
    license_window = tk.Toplevel(root)
    license_window.title("License Key Validation")
    license_window.geometry("400x200")
    license_window.resizable(False, False)

    tk.Label(license_window, text="Enter your license key:", font=("Verdana", 12)).pack(
        pady=10
    )
    entry_license = ttk.Entry(license_window, font=("Verdana", 12))
    entry_license.pack(pady=5, padx=10, fill=tk.X)

    ttk.Button(license_window, text="Submit", command=submit_key).pack(pady=10)

    license_window.protocol("WM_DELETE_WINDOW", root.destroy)


def fetch_all_definitions(word):
    """Fetch all available definitions and examples of a word using the dictionary API."""
    url = f"{DICTIONARY_API_URL}{word.lower()}"
    response = requests.get(url)
    if response.status_code != 200:
        return None
    data = response.json()

    # Extract all definitions and examples
    definitions = []
    for meaning in data[0]["meanings"]:
        part_of_speech = meaning.get("partOfSpeech", "Unknown")
        for definition in meaning["definitions"]:
            definitions.append(
                {
                    "partOfSpeech": part_of_speech,
                    "definition": definition["definition"],
                    "example": definition.get("example", "No example available"),
                }
            )
    return definitions


def view_definitions():
    """Display all definitions and examples for the selected word in a popup window."""
    selected_item = tree_vocabulary.selection()
    if not selected_item:
        messagebox.showwarning(
            "Selection Error", "Please select a word to view its definitions."
        )
        return

    word = tree_vocabulary.item(selected_item, "values")[0]
    definitions = fetch_all_definitions(word)

    if not definitions:
        messagebox.showerror(
            "Error", "Could not fetch definitions for the selected word."
        )
        return

    # Create a popup window
    definitions_window = tk.Toplevel(root)
    definitions_window.title(f"Definitions and Examples for '{word}'")
    definitions_window.geometry("600x500")
    definitions_window.resizable(True, True)

    ttk.Label(
        definitions_window,
        text=f"{word}",
        font=("Verdana", 14, "bold"),
    ).pack(pady=10)

    # Create a scrollable text widget to display definitions and examples
    frame = ttk.Frame(definitions_window)
    frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

    text_widget = tk.Text(frame, wrap=tk.WORD, font=("Verdana", 12))
    text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    scrollbar = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=text_widget.yview)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    text_widget.config(yscrollcommand=scrollbar.set)

    # Insert definitions and examples into the text widget
    for idx, definition in enumerate(definitions, start=1):
        part_of_speech = definition["partOfSpeech"]
        definition_text = definition["definition"]
        example_text = definition["example"]
        text_widget.insert(
            tk.END,
            f"{idx}. ({part_of_speech}) {definition_text}\n   Example: {example_text}\n\n",
        )

    text_widget.config(state=tk.DISABLED)  # Make the text widget read-only

    ttk.Button(
        definitions_window, text="Close", command=definitions_window.destroy
    ).pack(pady=10)


def show_license_status():
    """Display the license status linked to the current machine."""
    conn = mysql.connector.connect(**DB_CONFIG)
    cursor = conn.cursor()

    cursor.execute(
        """
        SELECT lk.license_key, lk.status, lk.expiry_date, lk.max_machines, COUNT(ma.machine_id) as activated_machines
        FROM license_keys lk
        JOIN machine_activations ma ON lk.license_key = ma.license_key
        WHERE ma.machine_id = %s
        GROUP BY lk.license_key
    """,
        (MACHINE_ID,),
    )
    row = cursor.fetchone()
    conn.close()

    if not row:
        messagebox.showinfo(
            "License Status", "No active license found for this machine."
        )
        return

    license_key, status, expiry_date, max_machines, activated_machines = row
    expiry_text = "No expiry" if expiry_date is None else expiry_date
    message = f"License Key: {license_key}\nStatus: {status}\nExpiry: {expiry_text}\nMax Machines: {max_machines}\nActivated Machines: {activated_machines}"
    messagebox.showinfo("License Status", message)


def show_about():
    """Display information about the application."""
    messagebox.showinfo(
        "About Vocabulary Manager",
        "Vocabulary Manager\nVersion 1.0\n\nAuthor: Kayden Lee\n© 2024",
    )


def check_db_connection():
    """Test the database connection."""
    try:
        conn = mysql.connector.connect(**DB_CONFIG)
        conn.close()
        messagebox.showinfo("Database Connection", "Database connection is active.")
    except mysql.connector.Error as e:
        messagebox.showerror("Database Connection", f"Database connection failed: {e}")


def adjust_column_widths(event):
    """Adjust column widths based on Treeview width."""
    total_width = tree_vocabulary.winfo_width()
    word_width = int(total_width * 0.2)
    meaning_width = int(total_width * 0.8)

    tree_vocabulary.column("Word", width=word_width)
    tree_vocabulary.column("Meaning", width=meaning_width)


def clear_selection(event=None):
    """Clear the current selection in the Treeview."""
    tree_vocabulary.selection_remove(tree_vocabulary.selection())


def on_double_click(event):
    """Handle double-click event on a row to display word definitions and examples."""
    global definition_window

    selected_item = tree_vocabulary.selection()
    if not selected_item:
        messagebox.showwarning(
            "Selection Error", "Please select a word to view its definitions."
        )
        return

    word = tree_vocabulary.item(selected_item, "values")[0]
    definitions = fetch_all_definitions(word)

    if not definitions:
        messagebox.showerror(
            "Error", "Could not fetch definitions for the selected word."
        )
        return

    # Check if a definition window is already open
    if definition_window and tk.Toplevel.winfo_exists(definition_window):
        # If the window exists, bring it to focus
        definition_window.focus()
        return

    # Create a new definition window
    definition_window = tk.Toplevel(root)
    definition_window.title(f"Definitions and Examples for '{word}'")
    definition_window.geometry("600x500")
    definition_window.resizable(True, True)

    ttk.Label(
        definition_window,
        text=f"{word}",
        font=("Verdana", 14, "bold"),
    ).pack(pady=10)

    # Create a scrollable text widget to display definitions and examples
    frame = ttk.Frame(definition_window)
    frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

    text_widget = tk.Text(frame, wrap=tk.WORD, font=("Verdana", 12))
    text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    scrollbar = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=text_widget.yview)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    text_widget.config(yscrollcommand=scrollbar.set)

    # Insert definitions and examples into the text widget
    for idx, definition in enumerate(definitions, start=1):
        part_of_speech = definition["partOfSpeech"]
        definition_text = definition["definition"]
        example_text = definition["example"]
        text_widget.insert(
            tk.END,
            f"{idx}. ({part_of_speech}) {definition_text}\n   Example: {example_text}\n\n",
        )

    text_widget.config(state=tk.DISABLED)  # Make the text widget read-only

    ttk.Button(
        definition_window, text="Close", command=lambda: close_window(definition_window)
    ).pack(pady=10)

    # Properly handle window close action
    definition_window.protocol(
        "WM_DELETE_WINDOW", lambda: close_window(definition_window)
    )


def close_window(window):
    """Close the definition window and reset the global reference."""
    global definition_window
    if window and tk.Toplevel.winfo_exists(window):
        window.destroy()
    definition_window = None


def clear_definition_window(window):
    """Clear the global reference when the definition window is closed."""
    global definition_window
    if window == definition_window:
        definition_window = None


def export_to_pdf():
    """Export the vocabulary list to a PDF file."""
    # Fetch vocabulary data
    conn = mysql.connector.connect(**DB_CONFIG)
    cursor = conn.cursor()
    cursor.execute("SELECT word, meaning FROM vocabulary")
    vocabulary = cursor.fetchall()
    conn.close()

    if not vocabulary:
        messagebox.showinfo("Export PDF", "No words found to export.")
        return

    # Create a PDF instance
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Arial", size=12)

    # Add a title
    pdf.set_font("Arial", style="B", size=14)
    pdf.cell(200, 10, txt="Vocabulary List", ln=True, align="C")
    pdf.ln(10)  # Add a line break

    # Add vocabulary data
    pdf.set_font("Arial", size=12)
    for word, meaning in vocabulary:
        pdf.cell(0, 10, txt=f"Word: {word}", ln=True)
        pdf.multi_cell(0, 10, txt=f"Meaning: {meaning}", align="L")
        pdf.ln(5)  # Add a small space between entries

    # Save the PDF
    try:
        pdf.output("Vocabulary_List.pdf")
        messagebox.showinfo(
            "Export PDF", "Vocabulary list has been exported to 'Vocabulary_List.pdf'."
        )
    except Exception as e:
        messagebox.showerror("Export PDF", f"An error occurred: {e}")


def export_to_xlsx():
    """Export the vocabulary list to an XLSX file."""
    # Fetch vocabulary data
    conn = mysql.connector.connect(**DB_CONFIG)
    cursor = conn.cursor()
    cursor.execute("SELECT word, meaning FROM vocabulary")
    vocabulary = cursor.fetchall()
    conn.close()

    if not vocabulary:
        messagebox.showinfo("Export XLSX", "No words found to export.")
        return

    # Create a new Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Vocabulary List"

    # Add headers to the Excel file
    ws.append(["Word", "Meaning"])

    # Add data rows
    for word, meaning in vocabulary:
        ws.append([word, meaning])

    # Save the workbook to a file
    try:
        file_path = os.path.join(os.getcwd(), f"Vocabulary_List_{TIMESTAMP}.xlsx")
        wb.save(file_path)
        messagebox.showinfo(
            "Export XLSX",
            f"Vocabulary list has been exported to '{file_path}'.",
        )
    except Exception as e:
        messagebox.showerror("Export XLSX", f"An error occurred: {e}")


url = "https://cdn.cloudservetechcentral.com/vocab-manager/32x32.ico"
response = requests.get(url)
with open("icon.ico", "wb") as file:
    file.write(response.content)


# Main Application Window
root = tk.Tk()
root.title("Vocabulary Manager")
root.state("zoomed")
root.geometry("900x600")
root.iconbitmap("icon.ico")
root.option_add("*Font", "Verdana 10")

# Menu Bar
menubar = tk.Menu(root)

# File Menu
file_menu = tk.Menu(menubar, tearoff=0)

export_menu = tk.Menu(file_menu, tearoff=0)
export_menu.add_command(label="Export to PDF", command=export_to_pdf)
export_menu.add_command(label="Export to XLSX", command=export_to_xlsx)

file_menu.add_cascade(label="Export", menu=export_menu)

menubar.add_cascade(label="File", menu=file_menu)

# License Menu
license_menu = tk.Menu(menubar, tearoff=0)
license_menu.add_command(label="Check Activation Status", command=show_license_status)
menubar.add_cascade(label="License", menu=license_menu)

# Database Tools Menu
db_menu = tk.Menu(menubar, tearoff=0)
db_menu.add_command(label="Check Database Connection", command=check_db_connection)
db_menu.add_command(label="Refresh Data", command=lambda: load_vocabulary())
menubar.add_cascade(label="Database", menu=db_menu)

# About Menu
about_menu = tk.Menu(menubar, tearoff=0)
about_menu.add_command(label="About App", command=show_about)
menubar.add_cascade(label="About", menu=about_menu)

root.config(menu=menubar)

# Configure Treeview
style = ttk.Style()
style.theme_use("clam")
style.configure("Treeview", font=("Verdana", 10), rowheight=30)
style.configure("Treeview.Heading", font=("Verdana", 12, "bold"))

columns = ("Word", "Meaning")
tree_vocabulary = ttk.Treeview(root, columns=columns, show="headings")
tree_vocabulary.heading("Word", text="Word")
tree_vocabulary.heading("Meaning", text="Meaning")
tree_vocabulary.column("Word", anchor="w")
tree_vocabulary.column("Meaning", anchor="w")
tree_vocabulary.pack(fill=tk.BOTH, expand=True, pady=10)

# Bind the resize event to adjust column widths dynamically
tree_vocabulary.bind("<Configure>", adjust_column_widths)

# Bind the Escape key to clear the selection
tree_vocabulary.bind("<Escape>", clear_selection)

# Bind the double-click event to show definitions
tree_vocabulary.bind("<Double-1>", on_double_click)

# Scrollbar for Treeview
scrollbar = ttk.Scrollbar(root, orient=tk.VERTICAL, command=tree_vocabulary.yview)
tree_vocabulary.configure(yscrollcommand=scrollbar.set)
scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

# Search Frame
frame_search = ttk.Frame(root, padding=10)
frame_search.pack(fill=tk.X)

ttk.Label(frame_search, text="Search:").pack(side=tk.LEFT, padx=5)
entry_search = ttk.Entry(frame_search)
entry_search.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
ttk.Button(frame_search, text="Search", command=on_search).pack(side=tk.LEFT, padx=5)

# Input Frame
frame_input = ttk.Frame(root, padding=10)
frame_input.pack(fill=tk.X)

ttk.Label(frame_input, text="Word/Phrase:").pack(side=tk.LEFT, padx=5)
entry_word = ttk.Entry(frame_input, width=30)
entry_word.pack(side=tk.LEFT, padx=5)
entry_word.bind("<Return>", add_word)

ttk.Label(frame_input, text="Meaning:").pack(side=tk.LEFT, padx=5)
entry_meaning = ttk.Entry(frame_input, width=50)
entry_meaning.pack(side=tk.LEFT, padx=5)
entry_meaning.bind("<Return>", add_word)

ttk.Button(frame_input, text="Add Word", command=add_word).pack(side=tk.LEFT, padx=5)
ttk.Button(frame_input, text="Edit Word", command=edit_word).pack(side=tk.LEFT, padx=5)
ttk.Button(frame_input, text="Delete Word", command=delete_word).pack(
    side=tk.LEFT, padx=5
)
ttk.Button(frame_input, text="View Definitions", command=view_definitions).pack(
    side=tk.LEFT, padx=5
)

# Initialize and Run Application
init_db()
show_license_key_entry()
load_vocabulary()

root.mainloop()
